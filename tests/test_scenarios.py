from datetime import date

import pandas as pd
import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from ecsa.db.models import Assignment, Center, Hall, School, Session as ExamSession
from ecsa.importer import import_dataframe
from ecsa.parameters.store import ParameterStore
from ecsa.reports import assignments_frame, attendance_sheets, export_scenario_excel, student_cards, utilization_frames
from ecsa.scenarios import RunRequest, approve_scenario, compare_scenarios, delete_scenario, list_scenarios, run_scenario
from ecsa.scenarios.service import ScenarioError
from ecsa.tools.sample_data import generate


@pytest.fixture()
def loaded(db, tmp_path):
    files = generate(tmp_path, governorate="Basra", students=600, schools=8, districts=3, seed=3)
    for kind in ("students", "subjects", "student_subjects", "schools"):
        rep = import_dataframe(db, kind, pd.read_csv(files[kind], dtype=str))
        assert rep.ok, rep.errors
    db.commit()
    return db


def _req(name="base", **over):
    return RunRequest(name, "Basra", 1, date(2026, 3, 1), param_overrides=over)


def test_end_to_end_run_seats_everyone(loaded):
    sc = run_scenario(loaded, _req(operating_days_per_round=10))
    assert sc.status == "completed", sc.kpi_summary
    k = sc.kpi_summary
    assert k["coverage"] == 1.0 and k["unassigned"] == 0 and k["validation"]["ok"]
    assert k["exam_cases"] == loaded.execute(select(func.count()).select_from(Assignment).where(Assignment.scenario_id == sc.scenario_id)).scalar()
    # snapshot carries every parameter incl. the override
    assert sc.params_snapshot["operating_days_per_round"] == "10"
    assert "reserve_center_ratio" in sc.params_snapshot
    # reserve centers are inactive and count = ceil(20% × main)
    centers = loaded.execute(select(Center).where(Center.scenario_id == sc.scenario_id)).scalars().all()
    main = [c for c in centers if c.category != "reserve"]
    reserve = [c for c in centers if c.category == "reserve"]
    assert all(c.activation_status == "inactive" for c in reserve)
    assert len(reserve) == -(-len(main) * 20 // 100) or sc.decision_log["decision"]["reserve_shortfall"] > 0
    # halls seeded from the school and safe capacity floored
    for c in main:
        halls = loaded.execute(select(Hall).where(Hall.center_id == c.center_id)).scalars().all()
        assert len(halls) == c.school.halls_count
        assert all(h.safe_capacity == int(h.capacity * 0.85) for h in halls)
    # low readiness schools never become centers
    assert all(c.school.readiness_score >= 60 for c in centers)
    assert sc.decision_log["ranking"] and sc.decision_log["schedule"]["slots"] == 40


def test_parameter_change_produces_a_different_distribution(loaded):
    a = run_scenario(loaded, _req("a", operating_days_per_round=10))
    b = run_scenario(loaded, _req("b", operating_days_per_round=10, safe_utilization_rate=0.6))
    assert b.kpi_summary["active_centers"] >= a.kpi_summary["active_centers"]
    assert b.kpi_summary["total_safe_seats"] != a.kpi_summary["total_safe_seats"]
    cmp = compare_scenarios(loaded, [a.scenario_id, b.scenario_id])
    assert cmp[0]["params_diff"] == {"safe_utilization_rate": "0.85"} and cmp[1]["params_diff"] == {"safe_utilization_rate": "0.6"}


def test_governorate_parameter_is_used(loaded):
    ParameterStore(loaded).set("sessions_per_day", 2, governorate="Basra", effective_from=date(2020, 1, 1))
    sc = run_scenario(loaded, _req(operating_days_per_round=20))
    assert sc.params_snapshot["sessions_per_day"] == "2"
    assert sc.decision_log["schedule"]["slots"] == 40


def test_include_exclude_schools(loaded):
    base = run_scenario(loaded, _req(operating_days_per_round=10))
    first = base.decision_log["decision"]["primary"][0]
    eligible_ids = [r["school_id"] for r in base.decision_log["ranking"] if r["eligible"]]
    last = eligible_ids[-1]
    sc = run_scenario(loaded, RunRequest("x", "Basra", 1, date(2026, 3, 1), {"operating_days_per_round": 10},
                                         include_schools=[last], exclude_schools=[first]))
    assert sc.decision_log["decision"]["primary"][0] == last
    assert first not in {r["school_id"] for r in sc.decision_log["ranking"]}


def test_approve_marks_schools_and_single_approved_per_round(loaded):
    a = run_scenario(loaded, _req("a", operating_days_per_round=10))
    b = run_scenario(loaded, _req("b", operating_days_per_round=10))
    approve_scenario(loaded, a.scenario_id)
    approve_scenario(loaded, b.scenario_id)
    loaded.expire_all()
    assert loaded.get(type(a), a.scenario_id).status == "completed"
    assert loaded.get(type(b), b.scenario_id).status == "approved"
    approved = {s.school_id for s in loaded.execute(select(School).where(School.is_approved)).scalars()}
    assert approved == {c.school_id for c in loaded.execute(select(Center).where(Center.scenario_id == b.scenario_id)).scalars()}


def test_delete_scenario_removes_children(loaded):
    sc = run_scenario(loaded, _req(operating_days_per_round=10))
    sid = sc.scenario_id
    delete_scenario(loaded, sid)
    assert not list_scenarios(loaded, "Basra")
    for model in (Assignment, ExamSession, Center):
        assert loaded.execute(select(func.count()).select_from(model).where(model.scenario_id == sid)).scalar() == 0


def test_run_without_data_raises(db):
    with pytest.raises(ScenarioError):
        run_scenario(db, RunRequest("x", "Nowhere", 1, date(2026, 1, 1)))


def test_too_few_slots_is_a_scenario_error(loaded):
    with pytest.raises(ScenarioError, match="scheduling failed"):
        run_scenario(loaded, _req(operating_days_per_round=1, sessions_per_day=1))


def test_shortfall_scenario_is_marked_failed_but_persisted(loaded):
    sc = run_scenario(loaded, _req(operating_days_per_round=2, safe_utilization_rate=0.05))
    assert sc.status == "failed" and sc.kpi_summary["unassigned"] > 0
    assert sc.decision_log["decision"]["capacity_shortfall"]


def test_reports_and_excel_export(loaded, tmp_path):
    sc = run_scenario(loaded, _req(operating_days_per_round=10))
    df = assignments_frame(loaded, sc.scenario_id)
    assert len(df) == sc.kpi_summary["assigned"]
    sheets = attendance_sheets(df)
    assert sheets and all((g["seat_no"].diff().dropna() == 1).all() for g in sheets.values())
    cards = student_cards(df)
    assert cards.groupby("student_id").size().max() <= 7
    util = utilization_frames(loaded, sc.scenario_id, df)
    assert (util["halls"]["utilization"] <= 1.0).all()
    out = export_scenario_excel(loaded, sc.scenario_id, tmp_path / "s.xlsx", attendance=True)
    wb = load_workbook(out, read_only=True)
    names = set(wb.sheetnames)
    assert {"Summary", "Parameters", "School_Ranking", "Centers", "Sessions", "Student_Cards", "Util_Centers"} <= names
    assert any(n.startswith("Att_") for n in names)
